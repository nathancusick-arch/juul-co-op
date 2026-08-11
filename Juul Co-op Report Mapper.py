import io
import re
from copy import copy
from datetime import date, datetime, time, timedelta
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


COLUMN_MAP = {
    "ID": "internal_id",
    "Retailer": "Retailer",
    "Hub": "Hub",
    "Location Name": "Location Name",
    "Premises Name": "site_name",
    "Address": "site_address_1",
    "Post Code": "site_post_code",
    "Date of visit": "date_of_visit_local",
    "Time of visit": "time_of_visit_local",
    "Site Code": "site_code",
    "Pass/Fail": "primary_result",
    "Were you able to successfully conduct this audit?": "Were you able to successfully conduct this audit?",
    "Abort Reason": "What was the reason for aborting this audit?",
    "Abort Category": None,
    "Fail Counter": None,
    "Pass After Fail": None,
    "Pass Counter": None,
    "Fail After Pass": None,
    "Please detail why you were unable to conduct this audit:": "Please detail why you were unable to conduct this audit:",
    "How long have you been a mystery shopper? (for this company, or another company)": None,
    "Please enter your age:": "Please enter your age:",
    "Please enter your gender:": "Please enter your gender:",
    "Did you have a beard at the time of the audit?": "Did you have a beard at the time of the audit?",
    "Were you wearing any facial cosmetic products at the time of the audit?": "Were you wearing any facial cosmetic products at the time of the audit?",
    "Did the store sell Juul products?": "Did the store sell Juul products?",
    "Where were the Juul products located in the store?": "Where were the Juul products located in the store?",
    "Did you see any non-Juul branded items that were labelled ''JUUL compatible pods\" in the store during your audit?": "Did you see any non-Juul branded items that were labelled ''JUUL compatible pods\" in the store during your audit?",
    "If so, please give details:": "If so, please give details:",
    "Did you see 'Challenge 25' signage in the store?": "Did you see 'Challenge 25' signage in the store?",
    "Was the signage JUUL branded?": "Was the signage JUUL branded?",
    "Please detail the store employee's name (if wearing a name badge). If there was no name badge please record an accurate description of the employee:": "Please detail the store employee's name (if wearing a name badge). If there was no name badge please record an accurate description of the employee:",
    "What was the gender of the employee who served you?": "What was the gender of the employee who served you?",
    "In which age group was the employee?": "In which age group was the employee?",
    "Were Juul pods available to purchase?": "Were Juul pods available to purchase?",
    "Please detail the product you attempted to purchase:": "Please detail the product you attempted to purchase:",
    "Did the person who served you ask for ID?": "Did the person who served you ask for ID?",
    "Please confirm that you did not present any ID:": "Please confirm that you did not present any ID:",
    "Did the store colleague allow you to purchase the restricted item without providing ID?": "Did the store colleague allow you to purchase the restricted item without providing ID?",
    "At what point were you asked for ID?": "At what point were you asked for ID?",
    "Were you wearing a protective face covering?": None,
    "Did the employee request your ID when you asked to purchase Juul pods with your face covering on or off?": None,
    "Did the employee who served you make eye contact with you?": "Did the employee who served you make eye contact with you?",
    "When was eye contact first made?": "When was eye contact first made?",
    "Were you given a receipt?": "Were you given a receipt?",
    "From the receipt, please enter any visible codes and employee name if any:": "From the receipt, please enter any visible codes and employee name if any:",
    "Did you see any JUUL branded adverts/posters visible from the outside of the store? If yes , please make sure you upload photo": "Did you see any JUUL branded adverts/posters visible from the outside of the store? If yes , please make sure you upload photo",
    "Was there anything about the interaction that you think JUUL should take note of?": "Was there anything about the interaction that you think JUUL should take note of?",
    "If so, please detail the interaction:": "If so, please detail the interaction:",
    "Month": "__KNIME_MONTH__",
    "Year": "__KNIME_YEAR__",
}

REQUIRED_EXPORT_COLUMNS = {
    "client_name",
    "internal_id",
    "site_code",
    "site_name",
    "site_address_1",
    "site_post_code",
    "date_of_visit_local",
    "time_of_visit_local",
    "primary_result",
}

LDM_SHEET = "LDM_All_Stores"
LDM_COLUMNS = ["Post Code", "HUB Number", "Location Name"]
REPORT_PREFIX = "The Co-operative weekly_"


def clean_postcode(value):
    """Return a postcode in uppercase outward/inward form for matching."""
    if value is None or pd.isna(value):
        return ""
    compact = "".join(str(value).upper().split())
    if len(compact) < 5:
        return ""
    return compact[:-3] + " " + compact[-3:]


def clean_text(value):
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def integer_if_possible(value):
    text = clean_text(value)
    if re.fullmatch(r"[+-]?\d+(?:\.0+)?", text):
        return int(float(text))
    return text


def most_recent_saturday(as_of_date):
    return as_of_date - timedelta(days=(as_of_date.weekday() - 5) % 7)


def parse_visit_time(value):
    text = clean_text(value)
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    return None


def read_audit_export(file_bytes):
    try:
        export = pd.read_csv(
            io.BytesIO(file_bytes),
            dtype=str,
            keep_default_na=False,
            encoding="utf-8-sig",
        )
    except UnicodeDecodeError:
        export = pd.read_csv(
            io.BytesIO(file_bytes),
            dtype=str,
            keep_default_na=False,
            encoding="cp1252",
        )

    missing = sorted(REQUIRED_EXPORT_COLUMNS - set(export.columns))
    if missing:
        raise ValueError(
            "The audit export is missing required column(s): " + ", ".join(missing)
        )
    return export


def read_ldm(file_bytes):
    try:
        ldm = pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=LDM_SHEET,
            dtype=str,
            keep_default_na=False,
        )
    except ValueError as exc:
        raise ValueError(
            f"The LDM must contain a sheet named '{LDM_SHEET}'."
        ) from exc

    missing = [column for column in LDM_COLUMNS if column not in ldm.columns]
    if missing:
        raise ValueError(
            "The LDM is missing required column(s): " + ", ".join(missing)
        )

    ldm = ldm[LDM_COLUMNS].copy()
    ldm["_PostCodeClean"] = ldm["Post Code"].map(clean_postcode)
    ldm = ldm[ldm["_PostCodeClean"].ne("")]
    return ldm


def build_ldm_lookup(ldm, required_postcodes):
    lookup = {}
    ambiguous = []

    for postcode in sorted(set(required_postcodes) - {""}):
        matches = ldm[ldm["_PostCodeClean"].eq(postcode)].copy()
        if matches.empty:
            continue

        active = matches[
            ~matches["Location Name"].str.contains(
                r"\(Not trading\)", case=False, na=False, regex=True
            )
        ]
        candidates = active if not active.empty else matches
        candidates = candidates.drop_duplicates(
            subset=["HUB Number", "Location Name"], keep="first"
        )

        if len(candidates) > 1:
            ambiguous.append(postcode)
            continue

        match = candidates.iloc[0]
        lookup[postcode] = (
            integer_if_possible(match["HUB Number"]),
            clean_text(match["Location Name"]),
        )

    if ambiguous:
        raise ValueError(
            "The LDM contains more than one active store for the following audit "
            "postcode(s), so Hub and Location Name cannot be selected safely: "
            + ", ".join(ambiguous)
        )

    return lookup


def find_report_sheet(workbook):
    expected_headers = list(COLUMN_MAP)
    for worksheet in workbook.worksheets:
        actual_headers = [
            clean_text(worksheet.cell(1, column).value)
            for column in range(1, len(expected_headers) + 1)
        ]
        if actual_headers == expected_headers:
            return worksheet

    raise ValueError(
        "The previous report does not contain a worksheet with the expected "
        "Juul Co-op report columns in row 1."
    )


def existing_report_ids(worksheet):
    ids = set()
    for row in range(2, worksheet.max_row + 1):
        value = clean_text(worksheet.cell(row, 1).value)
        if value:
            ids.add(value)
    return ids


def copy_row_format(worksheet, source_row, target_row, max_column):
    if source_row < 2:
        return

    for column in range(1, max_column + 1):
        source = worksheet.cell(source_row, column)
        target = worksheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format

    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel


def mapped_value(row, output_column, source_column):
    if source_column is None:
        return None
    if output_column == "Date of visit":
        parsed = row["_visit_date"]
        return parsed.to_pydatetime().replace(tzinfo=None)
    if output_column == "Time of visit":
        return row["_visit_time"]
    if output_column in {"Hub", "Please enter your age:", "Month", "Year"}:
        return integer_if_possible(row.get(source_column, ""))
    return clean_text(row.get(source_column, "")) or None


def extend_tables(worksheet, last_row, max_column):
    for table in worksheet.tables.values():
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        if min_row == 1 and min_col == 1 and max_col == max_column:
            table.ref = (
                f"A1:{get_column_letter(max_column)}{last_row}"
            )


def prepare_new_rows(export, ldm, worksheet, cutoff_date):
    coop = export[
        export["site_code"].astype(str).str.startswith("CoopTCG", na=False)
    ].copy()

    if coop.empty:
        return coop, []

    coop["_visit_date"] = pd.to_datetime(
        coop["date_of_visit_local"], dayfirst=True, errors="coerce"
    )
    bad_dates = coop[coop["_visit_date"].isna()]["internal_id"].map(clean_text)
    if not bad_dates.empty:
        raise ValueError(
            "These Co-op audits have an invalid date_of_visit_local: "
            + ", ".join(bad_dates.tolist())
        )

    coop = coop[coop["_visit_date"].dt.date.le(cutoff_date)].copy()
    if coop.empty:
        return coop, []

    coop["internal_id"] = coop["internal_id"].map(clean_text)
    coop = coop[coop["internal_id"].ne("")]
    coop = coop.drop_duplicates(subset=["internal_id"], keep="last")
    coop = coop[~coop["internal_id"].isin(existing_report_ids(worksheet))].copy()
    if coop.empty:
        return coop, []

    coop["_visit_time"] = coop["time_of_visit_local"].map(parse_visit_time)
    bad_times = coop[coop["_visit_time"].isna()]["internal_id"].tolist()
    if bad_times:
        raise ValueError(
            "These Co-op audits have an invalid time_of_visit_local: "
            + ", ".join(bad_times)
        )

    coop["_PostCodeClean"] = coop["site_post_code"].map(clean_postcode)
    ldm_lookup = build_ldm_lookup(ldm, coop["_PostCodeClean"].tolist())

    coop["Hub"] = coop["_PostCodeClean"].map(
        lambda postcode: ldm_lookup.get(postcode, ("NOT ON THE LIST", ""))[0]
    )
    coop["Location Name"] = coop["_PostCodeClean"].map(
        lambda postcode: ldm_lookup.get(postcode, ("", "NOT ON THE LIST"))[1]
    )
    unmatched = sorted(
        coop.loc[~coop["_PostCodeClean"].isin(ldm_lookup), "site_post_code"]
        .map(clean_text)
        .unique()
        .tolist()
    )

    coop["Retailer"] = coop["client_name"].replace(
        "Juul", "Co-operative Group Limited"
    )
    coop["primary_result"] = coop["primary_result"].map(clean_text).str.upper()
    coop["__KNIME_MONTH__"] = coop["_visit_date"].dt.month
    coop["__KNIME_YEAR__"] = coop["_visit_date"].dt.year
    coop["_sort_time"] = coop["_visit_time"].map(
        lambda value: (value.hour, value.minute, value.second)
    )
    coop = coop.sort_values(
        by=["_visit_date", "_sort_time", "internal_id"], kind="stable"
    )
    return coop, unmatched


def generate_report(audit_bytes, ldm_bytes, previous_report_bytes, as_of_date=None):
    if as_of_date is None:
        as_of_date = datetime.now(ZoneInfo("Europe/London")).date()
    elif isinstance(as_of_date, datetime):
        as_of_date = as_of_date.date()
    if not isinstance(as_of_date, date):
        raise TypeError("as_of_date must be a date or datetime")

    export = read_audit_export(audit_bytes)
    ldm = read_ldm(ldm_bytes)

    try:
        workbook = load_workbook(
            io.BytesIO(previous_report_bytes), data_only=False, keep_links=True
        )
    except Exception as exc:
        raise ValueError("The previous report is not a valid .xlsx workbook.") from exc

    worksheet = find_report_sheet(workbook)
    cutoff_date = most_recent_saturday(as_of_date)
    new_rows, unmatched_postcodes = prepare_new_rows(
        export, ldm, worksheet, cutoff_date
    )

    output_name = f"{REPORT_PREFIX}{as_of_date:%Y%m%d}.xlsx"
    if new_rows.empty:
        return previous_report_bytes, output_name, {
            "new_count": 0,
            "cutoff_date": cutoff_date,
            "unmatched_postcodes": [],
        }

    expected_headers = list(COLUMN_MAP)
    original_last_row = worksheet.max_row
    style_row = original_last_row if original_last_row >= 2 else 0

    for offset, (_, source_row) in enumerate(new_rows.iterrows(), start=1):
        target_row = original_last_row + offset
        copy_row_format(worksheet, style_row, target_row, len(expected_headers))

        for column_number, (output_column, source_column) in enumerate(
            COLUMN_MAP.items(), start=1
        ):
            worksheet.cell(target_row, column_number).value = mapped_value(
                source_row, output_column, source_column
            )

        worksheet.cell(target_row, 15).value = (
            f'=IF(AND($J{target_row}=$J{target_row}, $K{target_row}="FAIL"), '
            f'SUMPRODUCT(EXACT($J$2:$J{target_row}, $J{target_row})*1,'
            f'($K$2:$K{target_row}="FAIL")*1),0)'
        )
        worksheet.cell(target_row, 17).value = (
            f'=IF(AND($J{target_row}=$J{target_row}, $K{target_row}="PASS"), '
            f'SUMPRODUCT(EXACT($J$2:$J{target_row}, $J{target_row})*1,'
            f'($K$2:$K{target_row}="PASS")*1),0)'
        )

    final_row = worksheet.max_row
    extend_tables(worksheet, final_row, len(expected_headers))
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue(), output_name, {
        "new_count": len(new_rows),
        "cutoff_date": cutoff_date,
        "unmatched_postcodes": unmatched_postcodes,
    }


def main():
    st.title("Juul Co-op Report Generator")
    st.write(
        "Upload the audit export, the current Juul Co-op LDM, and the previous "
        "weekly report. The generator will remove audits already reported and "
        "add the new eligible Co-op audits to the workbook."
    )

    audit_file = st.file_uploader(
        "1. Upload audits_basic_data_export.csv", type=["csv"]
    )
    ldm_file = st.file_uploader(
        "2. Upload Juul Co-op LDM.xlsx", type=["xlsx"]
    )
    previous_report_file = st.file_uploader(
        "3. Upload the previous The Co-operative weekly report", type=["xlsx"]
    )

    all_uploaded = all([audit_file, ldm_file, previous_report_file])
    if st.button("Generate report", type="primary", disabled=not all_uploaded):
        try:
            report_bytes, output_name, details = generate_report(
                audit_file.getvalue(),
                ldm_file.getvalue(),
                previous_report_file.getvalue(),
            )
        except Exception as exc:
            st.error(str(exc))
        else:
            count = details["new_count"]
            audit_word = "audit" if count == 1 else "audits"
            st.success(
                f"Report generated with {count} new {audit_word}. "
                f"Audit cut-off: {details['cutoff_date']:%d/%m/%Y}."
            )
            if details["unmatched_postcodes"]:
                st.warning(
                    "No LDM match was found for: "
                    + ", ".join(details["unmatched_postcodes"])
                    + ". Hub and Location Name have been set to "
                    "'NOT ON THE LIST' for those audits."
                )

            st.download_button(
                "Download updated weekly report",
                data=report_bytes,
                file_name=output_name,
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )


if __name__ == "__main__":
    main()
